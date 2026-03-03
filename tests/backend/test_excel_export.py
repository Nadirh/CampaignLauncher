import uuid
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.campaign import BiddingStrategy, CampaignStatus
from app.models.keyword import MatchType
from app.schemas.pipeline import (
    AdGroupResponse,
    AdResponse,
    CampaignGenerateResponse,
    KeywordResponse,
)
from app.services.excel_export import (
    _format_path,
    generate_workbook,
)


@pytest.fixture(autouse=True)
def setup_db():
    """Override the async DB fixture -- these tests don't need a database."""
    yield

AG_ID = uuid.uuid4()
CAMPAIGN_ID = uuid.uuid4()


def _make_campaign(**overrides) -> CampaignGenerateResponse:
    defaults = {
        "id": CAMPAIGN_ID,
        "name": "Test Campaign",
        "landing_page_url": "https://example.com",
        "status": CampaignStatus.REVIEW,
        "bidding_strategy": BiddingStrategy.MANUAL_CPC,
        "daily_budget": 50.0,
        "match_types": ["phrase", "exact"],
        "negative_keywords": ["free", "cheap"],
        "bid_value": 2.50,
        "location_targeting": "US, UK",
        "created_at": datetime(2026, 1, 1, 0, 0, 0),
        "updated_at": datetime(2026, 1, 2, 0, 0, 0),
        "ad_groups": [
            AdGroupResponse(
                id=AG_ID,
                name="Ad Group 1",
                campaign_id=CAMPAIGN_ID,
                keywords=[
                    KeywordResponse(
                        id=uuid.uuid4(),
                        text="buy widgets",
                        match_type=MatchType.PHRASE,
                        bid=1.50,
                        ad_group_id=AG_ID,
                    ),
                    KeywordResponse(
                        id=uuid.uuid4(),
                        text="best widgets",
                        match_type=MatchType.EXACT,
                        bid=None,
                        ad_group_id=AG_ID,
                    ),
                ],
                ads=[
                    AdResponse(
                        id=uuid.uuid4(),
                        final_url="https://example.com",
                        headlines=[
                            {"text": "Buy Widgets", "position": 1, "trigger": "urgency"},
                            {"text": "Best Price", "position": 2, "trigger": "value proposition"},
                        ],
                        descriptions=[
                            {"text": "Shop now", "trigger": "call to action"},
                            {"text": "Free shipping"},
                        ],
                        path1="widgets",
                        path2="buy",
                        ad_group_id=AG_ID,
                    ),
                ],
            ),
        ],
    }
    defaults.update(overrides)
    return CampaignGenerateResponse(**defaults)


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_workbook_has_three_sheets():
    wb = _load(generate_workbook(_make_campaign()))
    assert wb.sheetnames == ["Summary", "Keywords", "Ads"]


def test_summary_sheet_values():
    campaign = _make_campaign()
    wb = _load(generate_workbook(campaign))
    ws = wb["Summary"]
    assert ws.cell(row=1, column=1).value == "Campaign Name"
    assert ws.cell(row=1, column=2).value == "Test Campaign"
    assert ws.cell(row=2, column=2).value == "https://example.com"
    assert ws.cell(row=3, column=2).value == "review"
    assert ws.cell(row=4, column=2).value == "manual_cpc"
    assert ws.cell(row=5, column=2).value == 50.0
    assert ws.cell(row=6, column=1).value == "Match Types"
    assert ws.cell(row=6, column=2).value == "phrase, exact"
    assert ws.cell(row=7, column=1).value == "Negative Keywords"
    assert ws.cell(row=7, column=2).value == "free, cheap"
    assert ws.cell(row=8, column=1).value == "Bid Value"
    assert ws.cell(row=8, column=2).value == 2.50
    assert ws.cell(row=9, column=1).value == "Location Targeting"
    assert ws.cell(row=9, column=2).value == "US, UK"
    assert ws.cell(row=10, column=2).value == 1  # ad groups
    assert ws.cell(row=11, column=2).value == 2  # keywords
    assert ws.cell(row=12, column=2).value == 1  # ads


def test_keywords_sheet_row_count():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Keywords"]
    # header + 2 keywords
    assert ws.max_row == 3


def test_keywords_sheet_values():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Keywords"]
    assert ws.cell(row=1, column=1).value == "Ad Group"
    assert ws.cell(row=2, column=1).value == "Ad Group 1"
    assert ws.cell(row=2, column=2).value == "buy widgets"
    assert ws.cell(row=2, column=3).value == "phrase"
    assert ws.cell(row=2, column=4).value == 1.50
    assert ws.cell(row=3, column=2).value == "best widgets"
    assert ws.cell(row=3, column=3).value == "exact"
    assert ws.cell(row=3, column=4) is not None


def test_ads_sheet_headers():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Ads"]
    assert ws.cell(row=1, column=1).value == "Ad Group"
    assert ws.cell(row=1, column=2).value == "Ad #"
    assert ws.cell(row=1, column=3).value == "Type"
    assert ws.cell(row=1, column=4).value == "#"
    assert ws.cell(row=1, column=5).value == "Copy"
    assert ws.cell(row=1, column=6).value == "Pin Position"
    assert ws.cell(row=1, column=7).value == "Trigger"
    assert ws.cell(row=1, column=8).value == "Length"
    assert ws.cell(row=1, column=9).value == "Final URL"
    assert ws.cell(row=1, column=10).value == "Path"


def test_ads_sheet_headline_rows():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Ads"]
    # Row 2: first headline
    assert ws.cell(row=2, column=1).value == "Ad Group 1"
    assert ws.cell(row=2, column=2).value == 1
    assert ws.cell(row=2, column=3).value == "Headline"
    assert ws.cell(row=2, column=4).value == 1
    assert ws.cell(row=2, column=5).value == "Buy Widgets"
    assert ws.cell(row=2, column=6).value == 1
    assert ws.cell(row=2, column=7).value == "urgency"
    assert ws.cell(row=2, column=8).value == len("Buy Widgets")
    assert ws.cell(row=2, column=9).value == "https://example.com"
    assert ws.cell(row=2, column=10).value == "/widgets/buy"
    # Row 3: second headline
    assert ws.cell(row=3, column=3).value == "Headline"
    assert ws.cell(row=3, column=4).value == 2
    assert ws.cell(row=3, column=5).value == "Best Price"
    assert ws.cell(row=3, column=6).value == 2
    assert ws.cell(row=3, column=7).value == "value proposition"
    assert ws.cell(row=3, column=8).value == len("Best Price")


def test_ads_sheet_description_rows():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Ads"]
    # Row 4: first description (after 2 headlines)
    assert ws.cell(row=4, column=3).value == "Description"
    assert ws.cell(row=4, column=4).value == 1
    assert ws.cell(row=4, column=5).value == "Shop now"
    assert ws.cell(row=4, column=6).value is None  # no pin position for descriptions
    assert ws.cell(row=4, column=7).value == "call to action"
    assert ws.cell(row=4, column=8).value == len("Shop now")
    # Row 5: second description (no trigger)
    assert ws.cell(row=5, column=3).value == "Description"
    assert ws.cell(row=5, column=4).value == 2
    assert ws.cell(row=5, column=5).value == "Free shipping"
    assert ws.cell(row=5, column=7).value is None
    assert ws.cell(row=5, column=8).value == len("Free shipping")


def test_ads_sheet_row_count():
    wb = _load(generate_workbook(_make_campaign()))
    ws = wb["Ads"]
    # header + 2 headlines + 2 descriptions = 5
    assert ws.max_row == 5


def test_empty_campaign():
    campaign = _make_campaign(ad_groups=[])
    wb = _load(generate_workbook(campaign))
    assert wb.sheetnames == ["Summary", "Keywords", "Ads"]
    assert wb["Keywords"].max_row == 1  # header only
    assert wb["Ads"].max_row == 1  # header only
    assert wb["Summary"].cell(row=10, column=2).value == 0


def test_format_path_both():
    assert _format_path("widgets", "buy") == "/widgets/buy"


def test_format_path_one():
    assert _format_path("widgets", None) == "/widgets"


def test_format_path_none():
    assert _format_path(None, None) == ""


def test_multiple_ad_groups():
    ag2_id = uuid.uuid4()
    campaign = _make_campaign(
        ad_groups=[
            AdGroupResponse(
                id=AG_ID,
                name="Group A",
                campaign_id=CAMPAIGN_ID,
                keywords=[
                    KeywordResponse(
                        id=uuid.uuid4(),
                        text="kw1",
                        match_type=MatchType.BROAD,
                        bid=None,
                        ad_group_id=AG_ID,
                    ),
                ],
                ads=[],
            ),
            AdGroupResponse(
                id=ag2_id,
                name="Group B",
                campaign_id=CAMPAIGN_ID,
                keywords=[
                    KeywordResponse(
                        id=uuid.uuid4(),
                        text="kw2",
                        match_type=MatchType.EXACT,
                        bid=2.0,
                        ad_group_id=ag2_id,
                    ),
                ],
                ads=[
                    AdResponse(
                        id=uuid.uuid4(),
                        final_url="https://example.com/b",
                        headlines=[{"text": "H1"}],
                        descriptions=[{"text": "D1"}],
                        path1=None,
                        path2=None,
                        ad_group_id=ag2_id,
                    ),
                ],
            ),
        ],
    )
    wb = _load(generate_workbook(campaign))
    assert wb["Keywords"].max_row == 3  # header + 2 keywords
    assert wb["Keywords"].cell(row=2, column=1).value == "Group A"
    assert wb["Keywords"].cell(row=3, column=1).value == "Group B"
    # header + 1 headline + 1 description = 3
    assert wb["Ads"].max_row == 3
    assert wb["Ads"].cell(row=2, column=1).value == "Group B"
    assert wb["Ads"].cell(row=2, column=3).value == "Headline"
    assert wb["Ads"].cell(row=2, column=5).value == "H1"
    assert wb["Ads"].cell(row=2, column=8).value == 2  # len("H1")
    assert wb["Ads"].cell(row=3, column=3).value == "Description"
    assert wb["Ads"].cell(row=3, column=5).value == "D1"
    # no path
    assert not wb["Ads"].cell(row=2, column=10).value


def test_returns_bytes():
    result = generate_workbook(_make_campaign())
    assert isinstance(result, bytes)
    assert len(result) > 0
