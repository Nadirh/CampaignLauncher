import uuid
from io import BytesIO

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ad import Ad
from app.models.ad_group import AdGroup
from app.models.campaign import Campaign, CampaignStatus
from app.models.keyword import Keyword, MatchType


async def _create_campaign_with_data(db: AsyncSession) -> Campaign:
    campaign = Campaign(
        name="Export Test",
        landing_page_url="https://example.com",
        status=CampaignStatus.REVIEW,
    )
    db.add(campaign)
    await db.flush()

    ag = AdGroup(name="Ad Group 1", campaign_id=campaign.id)
    db.add(ag)
    await db.flush()

    kw = Keyword(text="test keyword", match_type=MatchType.PHRASE, ad_group_id=ag.id)
    ad = Ad(
        final_url="https://example.com",
        headlines=[{"text": "Headline 1"}, {"text": "Headline 2"}],
        descriptions=[{"text": "Desc 1"}],
        path1="test",
        path2=None,
        ad_group_id=ag.id,
    )
    db.add_all([kw, ad])
    await db.commit()
    return campaign


async def test_export_returns_xlsx(
    client: AsyncClient, db_session: AsyncSession
) -> None:
    campaign = await _create_campaign_with_data(db_session)
    response = await client.get(f"/api/campaigns/{campaign.id}/export")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "attachment" in response.headers["content-disposition"]
    assert "Export_Test.xlsx" in response.headers["content-disposition"]


async def test_export_valid_workbook(
    client: AsyncClient, db_session: AsyncSession
) -> None:
    campaign = await _create_campaign_with_data(db_session)
    response = await client.get(f"/api/campaigns/{campaign.id}/export")
    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Summary", "Keywords", "Ads"]
    assert wb["Summary"].cell(row=1, column=2).value == "Export Test"
    assert wb["Keywords"].cell(row=2, column=2).value == "test keyword"
    assert wb["Ads"].cell(row=2, column=3).value == "Headline"
    assert wb["Ads"].cell(row=2, column=5).value == "Headline 1"


async def test_export_not_found(client: AsyncClient) -> None:
    fake_id = str(uuid.uuid4())
    response = await client.get(f"/api/campaigns/{fake_id}/export")
    assert response.status_code == 404


async def test_export_empty_campaign(
    client: AsyncClient, db_session: AsyncSession
) -> None:
    campaign = Campaign(
        name="Empty Campaign",
        landing_page_url="https://example.com",
        status=CampaignStatus.DRAFT,
    )
    db_session.add(campaign)
    await db_session.commit()
    response = await client.get(f"/api/campaigns/{campaign.id}/export")
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    assert wb["Keywords"].max_row == 1  # header only
    assert wb["Ads"].max_row == 1
