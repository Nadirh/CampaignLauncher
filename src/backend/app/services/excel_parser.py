from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.ads_script import (
    ParsedAd,
    ParsedCampaign,
    ParsedDescription,
    ParsedHeadline,
    ParsedKeyword,
    ParsedWorkbook,
)


class ExcelParseError(Exception):
    pass


def parse_workbook(file_bytes: bytes) -> ParsedWorkbook:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Cannot open file as Excel workbook: {exc}") from exc

    required_sheets = {"Summary", "Keywords", "Ads"}
    actual_sheets = set(wb.sheetnames)
    missing = required_sheets - actual_sheets
    if missing:
        raise ExcelParseError(f"Missing required sheets: {', '.join(sorted(missing))}")

    campaign = _parse_summary(wb["Summary"])
    keywords = _parse_keywords(wb["Keywords"])
    ads = _parse_ads(wb["Ads"])

    return ParsedWorkbook(campaign=campaign, keywords=keywords, ads=ads)


def _parse_summary(ws: Worksheet) -> ParsedCampaign:
    labels: dict[str, str | float | None] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        label, value = row
        if label:
            labels[str(label).strip()] = value

    def _get(key: str) -> str | float | None:
        return labels.get(key)

    name = _get("Campaign Name")
    if not name:
        raise ExcelParseError("Summary sheet missing 'Campaign Name'")

    landing_page_url = _get("Landing Page URL")
    if not landing_page_url:
        raise ExcelParseError("Summary sheet missing 'Landing Page URL'")

    bidding_strategy = _get("Bidding Strategy")
    if not bidding_strategy:
        raise ExcelParseError("Summary sheet missing 'Bidding Strategy'")

    daily_budget = _get("Daily Budget")
    if daily_budget is None:
        raise ExcelParseError("Summary sheet missing 'Daily Budget'")

    match_types_raw = _get("Match Types")
    match_types = None
    if match_types_raw:
        match_types = [m.strip() for m in str(match_types_raw).split(",") if m.strip()]

    neg_kw_raw = _get("Negative Keywords")
    negative_keywords = None
    if neg_kw_raw:
        negative_keywords = [k.strip() for k in str(neg_kw_raw).split(",") if k.strip()]

    bid_value = _get("Bid Value")
    location_targeting = _get("Location Targeting")

    return ParsedCampaign(
        name=str(name),
        landing_page_url=str(landing_page_url),
        bidding_strategy=str(bidding_strategy),
        daily_budget=float(daily_budget),
        match_types=match_types,
        negative_keywords=negative_keywords,
        bid_value=float(bid_value) if bid_value is not None else None,
        location_targeting=str(location_targeting) if location_targeting else None,
    )


def _parse_keywords(ws: Worksheet) -> list[ParsedKeyword]:
    keywords: list[ParsedKeyword] = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue
        ad_group = str(row[0]).strip()
        text = str(row[1]).strip() if row[1] else ""
        match_type = str(row[2]).strip() if row[2] else "phrase"
        bid = float(row[3]) if len(row) > 3 and row[3] is not None else None
        if text:
            keywords.append(ParsedKeyword(
                ad_group=ad_group,
                text=text,
                match_type=match_type,
                bid=bid,
            ))
    return keywords


def _parse_path(path_str: str | None) -> tuple[str | None, str | None]:
    if not path_str:
        return None, None
    stripped = path_str.strip().strip("/")
    if not stripped:
        return None, None
    parts = stripped.split("/", 1)
    path1 = parts[0] if parts[0] else None
    path2 = parts[1] if len(parts) > 1 and parts[1] else None
    return path1, path2


def _parse_ads(ws: Worksheet) -> list[ParsedAd]:
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Group rows by (ad_group, ad_number)
    ad_map: dict[tuple[str, int], dict] = {}

    for row in rows:
        if not row or not row[0]:
            continue
        ad_group = str(row[0]).strip()
        ad_number = int(row[1]) if row[1] is not None else 1
        row_type = str(row[2]).strip() if row[2] else ""
        copy_text = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        if not copy_text:
            continue

        key = (ad_group, ad_number)
        if key not in ad_map:
            final_url = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            path_str = str(row[9]).strip() if len(row) > 9 and row[9] else None
            path1, path2 = _parse_path(path_str)
            ad_map[key] = {
                "ad_group": ad_group,
                "ad_number": ad_number,
                "headlines": [],
                "descriptions": [],
                "final_url": final_url,
                "path1": path1,
                "path2": path2,
            }

        entry = ad_map[key]
        if row_type.lower() == "headline":
            pin = int(row[5]) if len(row) > 5 and row[5] is not None else None
            entry["headlines"].append(ParsedHeadline(text=copy_text, pin_position=pin))
        elif row_type.lower() == "description":
            entry["descriptions"].append(ParsedDescription(text=copy_text))

    ads: list[ParsedAd] = []
    for data in ad_map.values():
        ads.append(ParsedAd(
            ad_group=data["ad_group"],
            ad_number=data["ad_number"],
            headlines=data["headlines"],
            descriptions=data["descriptions"],
            final_url=data["final_url"],
            path1=data["path1"],
            path2=data["path2"],
        ))

    return ads
