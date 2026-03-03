from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.pipeline import CampaignGenerateResponse

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def generate_workbook(campaign: CampaignGenerateResponse) -> bytes:
    wb = Workbook()
    _write_summary_sheet(wb, campaign)
    _write_keywords_sheet(wb, campaign)
    _write_ads_sheet(wb, campaign)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary_sheet(wb: Workbook, campaign: CampaignGenerateResponse) -> None:
    ws = wb.active
    ws.title = "Summary"

    total_keywords = sum(len(ag.keywords) for ag in campaign.ad_groups)
    total_ads = sum(len(ag.ads) for ag in campaign.ad_groups)

    rows = [
        ("Campaign Name", campaign.name),
        ("Landing Page URL", campaign.landing_page_url),
        ("Status", campaign.status.value),
        ("Bidding Strategy", campaign.bidding_strategy.value),
        ("Daily Budget", campaign.daily_budget),
        ("Match Types", ", ".join(campaign.match_types) if campaign.match_types else None),
        ("Negative Keywords", ", ".join(campaign.negative_keywords) if campaign.negative_keywords else None),
        ("Bid Value", campaign.bid_value),
        ("Location Targeting", campaign.location_targeting),
        ("Ad Groups", len(campaign.ad_groups)),
        ("Total Keywords", total_keywords),
        ("Total Ads", total_ads),
        ("Created", campaign.created_at.isoformat()),
        ("Updated", campaign.updated_at.isoformat()),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_a.font = HEADER_FONT
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40


def _apply_header_row(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _write_keywords_sheet(wb: Workbook, campaign: CampaignGenerateResponse) -> None:
    ws = wb.create_sheet("Keywords")
    headers = ["Ad Group", "Keyword", "Match Type", "Bid"]
    _apply_header_row(ws, headers)

    row_idx = 2
    for ag in campaign.ad_groups:
        for kw in ag.keywords:
            ws.cell(row=row_idx, column=1, value=ag.name)
            ws.cell(row=row_idx, column=2, value=kw.text)
            ws.cell(row=row_idx, column=3, value=kw.match_type.value)
            ws.cell(row=row_idx, column=4, value=kw.bid)
            row_idx += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10


def _format_path(path1: str | None, path2: str | None) -> str:
    parts = [p for p in [path1, path2] if p]
    if not parts:
        return ""
    return "/" + "/".join(parts)


def _write_ads_sheet(wb: Workbook, campaign: CampaignGenerateResponse) -> None:
    ws = wb.create_sheet("Ads")
    headers = [
        "Ad Group", "Ad #", "Type", "#", "Copy",
        "Pin Position", "Trigger", "Length", "Final URL", "Path",
    ]
    _apply_header_row(ws, headers)

    row_idx = 2
    for ag in campaign.ad_groups:
        for ad_num, ad in enumerate(ag.ads, start=1):
            final_url = ad.final_url
            path = _format_path(ad.path1, ad.path2)

            for i, h in enumerate(ad.headlines):
                text = h["text"]
                ws.cell(row=row_idx, column=1, value=ag.name)
                ws.cell(row=row_idx, column=2, value=ad_num)
                ws.cell(row=row_idx, column=3, value="Headline")
                ws.cell(row=row_idx, column=4, value=i + 1)
                ws.cell(row=row_idx, column=5, value=text)
                pos = h.get("position")
                ws.cell(row=row_idx, column=6, value=pos)
                ws.cell(row=row_idx, column=7, value=h.get("trigger"))
                ws.cell(row=row_idx, column=8, value=len(text))
                ws.cell(row=row_idx, column=9, value=final_url)
                ws.cell(row=row_idx, column=10, value=path)
                row_idx += 1

            for i, d in enumerate(ad.descriptions):
                text = d["text"]
                ws.cell(row=row_idx, column=1, value=ag.name)
                ws.cell(row=row_idx, column=2, value=ad_num)
                ws.cell(row=row_idx, column=3, value="Description")
                ws.cell(row=row_idx, column=4, value=i + 1)
                ws.cell(row=row_idx, column=5, value=text)
                ws.cell(row=row_idx, column=7, value=d.get("trigger"))
                ws.cell(row=row_idx, column=8, value=len(text))
                ws.cell(row=row_idx, column=9, value=final_url)
                ws.cell(row=row_idx, column=10, value=path)
                row_idx += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 15
